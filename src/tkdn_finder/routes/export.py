# src/tkdn_finder/routes/export.py
"""Excel export route."""

from __future__ import annotations

import io
import logging
from datetime import date, datetime

from fastapi import APIRouter
from fastapi.responses import StreamingResponse

from ..config import get_settings
from ..constants import (
    SEARCH_RESULT_LIMIT_MAX,
    TKDN_DEFAULT_MIN_FILTER,
    VALIDITY_EXPIRING_SOON_DAYS,
)
from ..db import get_connection, get_last_refresh_ts
from ..models import CertificateRow
from ..search import search as do_search

logger = logging.getLogger(__name__)
router = APIRouter()


@router.get("/export.xlsx")
async def export_excel(
    q: str = "",
    tkdn_min: float = TKDN_DEFAULT_MIN_FILTER,
    validity_only: bool = False,
    kbli: str = "",
    year: str = "",
) -> StreamingResponse:
    """Export search results as an Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl not installed") from exc

    year_int: int | None = int(year) if year.strip() else None
    kbli_str: str | None = kbli.strip() or None

    settings = get_settings()
    conn = get_connection(settings.get_db_path())
    try:
        result = do_search(
            conn=conn,
            query=q,
            tkdn_min=tkdn_min,
            validity_only=validity_only,
            kbli=kbli_str,
            year=year_int,
            limit=SEARCH_RESULT_LIMIT_MAX,
            offset=0,
        )
        last_refresh_ts = get_last_refresh_ts(conn)
    finally:
        conn.close()

    # Format last_refresh in WIB for the audit trail
    from datetime import timedelta, timezone
    WIB = timezone(timedelta(hours=7))
    last_refresh_wib = "—"
    if last_refresh_ts:
        try:
            dt = datetime.fromisoformat(last_refresh_ts.replace(" ", "T", 1))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            last_refresh_wib = dt.astimezone(WIB).strftime("%d %b %Y %H:%M WIB")
        except ValueError:
            last_refresh_wib = last_refresh_ts[:19]

    today = date.today()
    rows = [CertificateRow.from_row(r, today) for r in result["results"]]

    wb = openpyxl.Workbook()

    # --- Info sheet ---
    ws_info = wb.active
    ws_info.title = "Info"
    ws_info.append(["TKDN Finder Export"])
    ws_info.append(["Diekspor pada", datetime.now(WIB).strftime("%d %b %Y %H:%M WIB")])
    ws_info.append(["Data P3DN terakhir diperbarui", last_refresh_wib])
    ws_info.append(["Query", q or "(none)"])
    ws_info.append(["TKDN Min", tkdn_min])
    ws_info.append(["Validity Only", str(validity_only)])
    ws_info.append(["KBLI Filter", kbli or "(none)"])
    ws_info.append(["Year Filter", str(year) if year else "(none)"])
    ws_info.append(["Total Results", result["total"]])
    ws_info.append(["Exported Rows", len(rows)])

    # --- Data sheet ---
    ws_data = wb.create_sheet("Data TKDN")
    headers = [
        "No",
        "Nama Perusahaan",
        "Nama Produk",
        "Spesifikasi",
        "Merek",
        "Tipe",
        "Nilai TKDN (%)",
        "Kode HS",
        "KBLI",
        "Kelompok Barang",
        "Alamat",
        "Provinsi",
        "Masa Berlaku Akhir",
        "Tahun Sumber",
        "Status Validitas",
    ]
    ws_data.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws_data[1]:
        cell.fill = header_fill
        cell.font = header_font

    for idx, cert in enumerate(rows, start=1):
        if cert.masa_berlaku_akhir is None:
            validity = "Unknown"
        elif cert.masa_berlaku_akhir < today:
            validity = "Expired"
        elif (cert.masa_berlaku_akhir - today).days <= VALIDITY_EXPIRING_SOON_DAYS:
            validity = "Expiring Soon"
        else:
            validity = "Valid"

        ws_data.append([
            idx,
            cert.nama_perusahaan,
            cert.nama_produk,
            cert.spesifikasi,
            cert.merek or "",
            cert.tipe or "",
            cert.nilai_tkdn,
            cert.kode_hs or "",
            cert.kbli or "",
            cert.kelompok_barang or "",
            cert.alamat or "",
            cert.provinsi or "",
            cert.masa_berlaku_akhir.isoformat() if cert.masa_berlaku_akhir else "",
            cert.tahun_sumber,
            validity,
        ])

    # Auto-size columns approximately
    for col in ws_data.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws_data.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # Write to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"tkdn_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
